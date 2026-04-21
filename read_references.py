import os
import sys
import zipfile
import xml.etree.ElementTree as ET

# Try to import pdfplumber for PDF reading
try:
    import pdfplumber
    PDF_READER = 'pdfplumber'
except ImportError:
    try:
        import PyPDF2
        PDF_READER = 'pypdf2'
    except ImportError:
        PDF_READER = None

# Try to import openpyxl for xlsx reading
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

REF_DIR = r"c:\Users\PC\Desktop\SDA PROJECT\reference"
OUTPUT_FILE = r"c:\Users\PC\Desktop\SDA PROJECT\reference_notes.txt"

def read_pdf(filepath):
    text = ""
    if PDF_READER == 'pdfplumber':
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    text += f"\n--- Page {i+1} ---\n{page_text}"
    elif PDF_READER == 'pypdf2':
        with open(filepath, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text:
                    text += f"\n--- Page {i+1} ---\n{page_text}"
    else:
        text = "[ERROR: No PDF reader available. Install pdfplumber or PyPDF2]"
    return text

def read_xlsx(filepath):
    if not XLSX_AVAILABLE:
        return "[ERROR: openpyxl not available. Install it to read xlsx files]"
    wb = openpyxl.load_workbook(filepath)
    text = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text += f"\n=== Sheet: {sheet_name} ===\n"
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if any(r.strip() for r in row_data):
                text += " | ".join(row_data) + "\n"
    return text

def read_drawio(filepath):
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
        text = f"[DrawIO file: {os.path.basename(filepath)}]\n"
        # Extract labels from mxCell elements
        for elem in root.iter():
            label = elem.get('label', '')
            value = elem.get('value', '')
            if label and label.strip():
                text += f"  Label: {label.strip()}\n"
            elif value and value.strip():
                text += f"  Value: {value.strip()}\n"
        return text
    except Exception as e:
        return f"[ERROR reading drawio: {e}]"

def main():
    output_lines = []
    
    # Check available libraries
    output_lines.append(f"PDF Reader: {PDF_READER}")
    output_lines.append(f"XLSX Available: {XLSX_AVAILABLE}")
    output_lines.append("="*80)
    
    # Read all files in reference directory
    for filename in os.listdir(REF_DIR):
        filepath = os.path.join(REF_DIR, filename)
        
        if os.path.isdir(filepath):
            # Handle unzipped_D5 directory with drawio files
            output_lines.append(f"\n{'='*80}")
            output_lines.append(f"DIRECTORY: {filename}")
            output_lines.append('='*80)
            for sub_filename in os.listdir(filepath):
                sub_filepath = os.path.join(filepath, sub_filename)
                if sub_filename.endswith('.drawio'):
                    output_lines.append(f"\n--- FILE: {sub_filename} ---")
                    output_lines.append(read_drawio(sub_filepath))
                elif sub_filename.endswith('.pdf'):
                    output_lines.append(f"\n--- FILE: {sub_filename} ---")
                    output_lines.append(read_pdf(sub_filepath))
            continue
        
        output_lines.append(f"\n{'='*80}")
        output_lines.append(f"FILE: {filename}")
        output_lines.append('='*80)
        
        if filename.endswith('.pdf'):
            output_lines.append(read_pdf(filepath))
        elif filename.endswith('.xlsx'):
            output_lines.append(read_xlsx(filepath))
        elif filename.endswith('.zip'):
            output_lines.append(f"[ZIP file - contents already extracted to unzipped_D5]")
    
    full_text = "\n".join(output_lines)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(full_text)
    
    print(f"Done! Output written to: {OUTPUT_FILE}")
    print(f"Total characters: {len(full_text)}")

if __name__ == "__main__":
    main()
